from __future__ import annotations

from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from fund_evaluation_tool.app_logic import build_legacy_analysis
from fund_evaluation_tool.export import export_legacy_report_to_excel, export_to_excel

FIXTURE = Path(__file__).parent / "fixtures" / "legacy_annual_returns.csv"


def test_export_to_excel_preserves_monthly_workbook_shape():
    output = BytesIO()

    export_to_excel(
        {"Fund_A": {"annualised_return": 0.12, "sharpe_ratio": 1.1}},
        output=output,
    )

    output.seek(0)
    workbook = load_workbook(output)

    assert workbook.sheetnames == ["Metrics"]
    assert workbook["Metrics"]["A1"].value == "Fund"
    assert workbook["Metrics"]["B1"].value == "Annualised Return"


def test_export_legacy_report_writes_comparison_raw_data_and_assumptions():
    legacy_result = build_legacy_analysis(FIXTURE, risk_free_rate=0.04)
    output = BytesIO()

    export_legacy_report_to_excel(
        legacy_result["all_metrics"],
        raw_data_df=legacy_result["raw_data_df"],
        output=output,
        comparison_df=legacy_result["benchmark_comparison_df"],
        assumptions=legacy_result["assumptions_df"],
    )

    output.seek(0)
    workbook = load_workbook(output)

    assert workbook.sheetnames == ["Comparison", "Raw Data", "Assumptions"]

    comparison_headers = [cell.value for cell in workbook["Comparison"][1]]
    assert comparison_headers[:4] == ["Fund", "Fund Cagr", "Benchmark Cagr", "Excess Cagr"]
    assert "Fund Ips Compliant" in comparison_headers

    raw_headers = [cell.value for cell in workbook["Raw Data"][1]]
    assert raw_headers[:6] == [
        "Fund",
        "Year",
        "Fund Return",
        "Spx Return",
        "Is Partial Year",
        "Months In Period",
    ]
    assert workbook["Raw Data"]["A2"].value == "HCI"

    assumptions_headers = [cell.value for cell in workbook["Assumptions"][1]]
    assert assumptions_headers == ["Assumption", "Value", "Status", "Source"]
    values = [cell.value for cell in workbook["Assumptions"]["A"] if cell.row > 1]
    assert {
        "CPI assumption",
        "Risk-free rate",
        "Partial-year handling",
        "Fee treatment",
    }.issubset(set(values))
    assert workbook["Assumptions"]["B2"].value == 0.03
